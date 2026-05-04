"""
Upload catalogue Excel/CSV + gestion des produits
"""

import uuid
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from loguru import logger
from pydantic import BaseModel
from typing import Optional

from app.db.database import get_db
from app.db import crud
from app.db.models import Tenant
from app.auth.dependencies import get_current_tenant
from app.rag.pg_retriever import PgVectorRetriever

router = APIRouter(prefix="/api/tenants", tags=["Catalog"])


# ─── Schemas ───────────────────────────────────────────────

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    price: Optional[str] = None
    category: Optional[str] = None
    sizes: Optional[str] = None
    colors: Optional[str] = None
    stock_status: Optional[str] = None


# ─── Column detection ─────────────────────────────────────

COLUMN_MAPPING = {
    "name": ["nom", "name", "produit", "product", "titre", "title", "designation", "article"],
    "description": ["description", "desc", "details", "detail"],
    "price": ["prix", "price", "tarif", "cout", "cost", "montant"],
    "category": ["categorie", "category", "cat", "type", "famille"],
    "sizes": ["taille", "tailles", "size", "sizes", "pointure"],
    "colors": ["couleur", "couleurs", "color", "colors"],
    "stock_status": ["stock", "disponibilite", "dispo", "status", "statut", "availability"],
    "image_url": ["image", "image_url", "photo", "photo_url", "img", "url_image", "lien_image"],
}


def _detect_columns(headers: list[str]) -> dict[str, int]:
    """Detecte automatiquement le mapping colonnes → champs"""
    mapping = {}
    headers_lower = [h.strip().lower() for h in headers]

    for field, aliases in COLUMN_MAPPING.items():
        for i, header in enumerate(headers_lower):
            if header in aliases:
                mapping[field] = i
                break

    return mapping


def _product_to_text(product_data: dict) -> str:
    """Convertit un produit en texte enrichi pour l'embedding.

    Strategie pour maximiser le retrieval :
    1. Section structuree (champs cles) — capture les recherches precises
    2. Phrase naturelle resumee (FR + MG) — capture les questions conversationnelles
       en francais et malgache, langues principales des clients Mada.
    3. Synonymes implicites via mots-cles ajoutes (ex: "disponible/dispo/mbola misy")
    """
    name = (product_data.get("name") or "").strip()
    category = (product_data.get("category") or "").strip()
    description = (product_data.get("description") or "").strip()
    price = (product_data.get("price") or "").strip()
    sizes = (product_data.get("sizes") or "").strip()
    colors = (product_data.get("colors") or "").strip()
    stock = (product_data.get("stock_status") or "").strip()

    parts = []

    # --- Section structuree (matching precis sur champs) ---
    if name:
        parts.append(f"Produit: {name}")
    if category:
        parts.append(f"Categorie: {category}")
    if description:
        parts.append(f"Description: {description}")
    if price:
        parts.append(f"Prix: {price}")
    if sizes:
        parts.append(f"Tailles disponibles: {sizes}")
    if colors:
        parts.append(f"Couleurs disponibles: {colors}")
    if stock:
        parts.append(f"Disponibilite: {stock}")

    # --- Phrase naturelle FR + MG (matching conversationnel) ---
    # Genere une phrase fluide qui matche les questions clients typiques
    # ("c'est combien la X ?", "manana X ve ?", "y a t-il des X ?")
    natural_fr_parts = []
    natural_mg_parts = []

    if name:
        # FR
        natural_fr_parts.append(f"Nous vendons {name}")
        if category:
            natural_fr_parts.append(f"dans la categorie {category}")
        if price:
            natural_fr_parts.append(f"au prix de {price}")
        if stock:
            stock_lower = stock.lower()
            if "dispo" in stock_lower or "stock" in stock_lower:
                natural_fr_parts.append("c'est disponible")
            elif "rupture" in stock_lower:
                natural_fr_parts.append("actuellement en rupture de stock")

        # MG (malgache)
        natural_mg_parts.append(f"Manana {name} izahay")
        if price:
            natural_mg_parts.append(f"mitentina {price}")
        if stock:
            stock_lower = stock.lower()
            if "dispo" in stock_lower or "stock" in stock_lower:
                natural_mg_parts.append("mbola misy")
            elif "rupture" in stock_lower:
                natural_mg_parts.append("tsy misy intsony amin'izao fotoana izao")

    if natural_fr_parts:
        parts.append(". ".join(natural_fr_parts) + ".")
    if natural_mg_parts:
        parts.append(". ".join(natural_mg_parts) + ".")

    # --- Tags / mots-cles supplementaires ---
    keywords = []
    if name:
        keywords.append(name.lower())
    if category:
        keywords.append(category.lower())
    if colors:
        # Chaque couleur indexee individuellement pour matcher "casquette noir"
        for c in colors.split(","):
            c = c.strip().lower()
            if c:
                keywords.append(c)
    if keywords:
        parts.append(f"Mots-cles: {', '.join(keywords)}")

    return "\n".join(parts)


def _catalog_summary_text(products_data: list[dict]) -> str:
    """Genere un document 'resume catalogue' indexe en plus des produits.

    Permet aux questions du type 'C'est quoi votre catalogue ?' / 'Inona ny vidiana ?'
    / 'Vous vendez quoi ?' de matcher avec un score eleve, sinon ces queries
    generales tombent toutes en confidence=none car aucun produit individuel
    ne contient ces mots.
    """
    names = [p.get("name", "").strip() for p in products_data if p.get("name")]
    categories = sorted({(p.get("category") or "").strip() for p in products_data if p.get("category")})

    parts = []
    parts.append("Voici le catalogue complet de notre boutique.")

    if categories:
        parts.append(f"Nous vendons dans ces categories: {', '.join(categories)}.")
        parts.append(f"Mivarotra eo amin'ireto sokajy ireto izahay: {', '.join(categories)}.")

    if names:
        listing_fr = ", ".join(names)
        parts.append(f"La liste de tous nos produits: {listing_fr}.")
        parts.append(f"Ireto avy ny vokatra rehetra azonao vidiana: {listing_fr}.")

    parts.append(
        "Les clients peuvent poser des questions sur le catalogue, les produits, "
        "les prix, la disponibilite, les couleurs, les tailles, les categories, "
        "ou demander la liste de ce que nous vendons."
    )
    parts.append(
        "Ny mpanjifa dia afaka manontany momba ny vokatra, ny vidiny, "
        "ny mbola misy, ny loko, ny habe, ny sokajy, na manontany ny lisitry "
        "ny vokatra amidinay."
    )

    return "\n".join(parts)


# ─── Endpoints ─────────────────────────────────────────────

@router.post("/{tenant_id}/upload-catalog")
async def upload_catalog(
    tenant_id: str,
    file: UploadFile = File(...),
    tenant: Tenant = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db),
):
    """Upload un fichier Excel/CSV, parse les produits, genere les embeddings"""
    # Verifier que le tenant_id correspond
    if str(tenant.id) != tenant_id:
        raise HTTPException(status_code=403, detail="Acces refuse")

    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Format supporte: .xlsx ou .csv")

    content = await file.read()
    products_data = []

    try:
        if filename.endswith(".xlsx"):
            products_data = _parse_xlsx(content)
        else:
            products_data = _parse_csv(content)
    except Exception as e:
        logger.error(f"Erreur parsing fichier: {e}")
        raise HTTPException(status_code=400, detail=f"Erreur lecture fichier: {str(e)}")

    if not products_data:
        raise HTTPException(status_code=400, detail="Aucun produit trouve dans le fichier")

    # Supprimer les anciens produits et embeddings
    await crud.delete_tenant_products(db, tenant.id)
    await crud.delete_tenant_embeddings(db, tenant.id)

    # Inserer les nouveaux produits
    await crud.create_products(db, tenant.id, products_data)

    # Generer les embeddings : 1 doc par produit + 1 doc "resume catalogue"
    texts = [_product_to_text(p) for p in products_data]
    metadatas = [
        {
            "source": "catalog",
            "product_name": p.get("name", ""),
            "image_url": p.get("image_url", "") or "",
        }
        for p in products_data
    ]

    # Ajout du document resume catalogue (matche les queries generiques type
    # "c'est quoi votre catalogue ?" qui ne matchent aucun produit individuel)
    texts.append(_catalog_summary_text(products_data))
    metadatas.append({"source": "catalog_summary", "product_name": "", "image_url": ""})

    retriever = PgVectorRetriever(tenant_id=tenant.id, db=db)
    await retriever.add_documents(texts, metadatas)

    # Log l'upload
    await crud.create_upload(db, tenant.id, file.filename, len(products_data))

    logger.info(f"Catalogue uploade pour {tenant.page_name}: {len(products_data)} produits")

    return {
        "status": "success",
        "products_count": len(products_data),
        "embeddings_count": len(texts),
        "filename": file.filename,
    }


@router.get("/{tenant_id}/products")
async def list_products(
    tenant_id: str,
    tenant: Tenant = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db),
):
    """Liste les produits du tenant"""
    if str(tenant.id) != tenant_id:
        raise HTTPException(status_code=403, detail="Acces refuse")

    products = await crud.get_products(db, tenant.id)
    return [
        {
            "id": str(p.id),
            "name": p.name,
            "description": p.description,
            "price": p.price,
            "category": p.category,
            "sizes": p.sizes,
            "colors": p.colors,
            "stock_status": p.stock_status,
            "image_url": p.image_url,
        }
        for p in products
    ]


@router.put("/{tenant_id}/products/{product_id}")
async def update_product(
    tenant_id: str,
    product_id: str,
    data: ProductUpdate,
    tenant: Tenant = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db),
):
    """Modifie un produit"""
    if str(tenant.id) != tenant_id:
        raise HTTPException(status_code=403, detail="Acces refuse")

    product = await crud.get_product_by_id(db, uuid.UUID(product_id))
    if not product or product.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Produit non trouve")

    updates = data.model_dump(exclude_unset=True)
    updated = await crud.update_product(db, product, **updates)

    return {
        "id": str(updated.id),
        "name": updated.name,
        "description": updated.description,
        "price": updated.price,
        "category": updated.category,
    }


@router.delete("/{tenant_id}/products/{product_id}")
async def delete_product(
    tenant_id: str,
    product_id: str,
    tenant: Tenant = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db),
):
    """Supprime un produit"""
    if str(tenant.id) != tenant_id:
        raise HTTPException(status_code=403, detail="Acces refuse")

    product = await crud.get_product_by_id(db, uuid.UUID(product_id))
    if not product or product.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Produit non trouve")

    await crud.delete_product(db, product)
    return {"status": "deleted"}


@router.post("/{tenant_id}/reindex")
async def reindex_products(
    tenant_id: str,
    tenant: Tenant = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db),
):
    """Re-genere tous les embeddings a partir des produits"""
    if str(tenant.id) != tenant_id:
        raise HTTPException(status_code=403, detail="Acces refuse")

    # Supprimer les anciens embeddings
    await crud.delete_tenant_embeddings(db, tenant.id)

    # Recharger les produits
    products = await crud.get_products(db, tenant.id)
    if not products:
        return {"status": "no_products", "embeddings_count": 0}

    products_dicts = [{
        "name": p.name, "description": p.description, "price": p.price,
        "category": p.category, "sizes": p.sizes, "colors": p.colors,
        "stock_status": p.stock_status,
        "image_url": p.image_url,
    } for p in products]

    texts = [_product_to_text(pd) for pd in products_dicts]
    metadatas = [
        {
            "source": "catalog",
            "product_name": p.name,
            "image_url": p.image_url or "",
        }
        for p in products
    ]

    # Document resume catalogue
    texts.append(_catalog_summary_text(products_dicts))
    metadatas.append({"source": "catalog_summary", "product_name": "", "image_url": ""})

    retriever = PgVectorRetriever(tenant_id=tenant.id, db=db)
    await retriever.add_documents(texts, metadatas)

    return {"status": "reindexed", "embeddings_count": len(texts)}


# ─── Parsing helpers ──────────────────────────────────────

def _parse_xlsx(content: bytes) -> list[dict]:
    """Parse un fichier Excel"""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h) if h else "" for h in rows[0]]
    col_map = _detect_columns(headers)

    if "name" not in col_map:
        # Fallback: premiere colonne = nom
        col_map["name"] = 0

    products = []
    for row in rows[1:]:
        if not row or not any(row):
            continue

        product = {}
        for field, col_idx in col_map.items():
            if col_idx < len(row) and row[col_idx] is not None:
                product[field] = str(row[col_idx]).strip()
            else:
                product[field] = ""

        if product.get("name"):
            products.append(product)

    wb.close()
    return products


def _parse_csv(content: bytes) -> list[dict]:
    """Parse un fichier CSV"""
    import csv

    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if len(rows) < 2:
        return []

    headers = rows[0]
    col_map = _detect_columns(headers)

    if "name" not in col_map:
        col_map["name"] = 0

    products = []
    for row in rows[1:]:
        if not row or not any(row):
            continue

        product = {}
        for field, col_idx in col_map.items():
            if col_idx < len(row):
                product[field] = row[col_idx].strip()
            else:
                product[field] = ""

        if product.get("name"):
            products.append(product)

    return products
